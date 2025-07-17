from django.shortcuts import render
from django.views.decorators.csrf import csrf_protect

from django.contrib.auth.decorators import login_required
from .models import EmployeeMaster, ClientMaster
from django.shortcuts import redirect
# from .forms import CustomUserCreationForm
from django.contrib import messages
from django.contrib.auth import logout as auth_logout


# 
# def home(request):
#     return render(request, 'home.html')



@login_required
@csrf_protect
def home(request):
    return render(request, 'home.html')

@login_required
@csrf_protect
def logout(request):
    auth_logout(request)  # log out the user
    return redirect('login')  # redirect to login page



from django.shortcuts import render
from .models import EmployeeMaster, ClientMaster, Users, Planning
from django.contrib.auth.models import User


from django.contrib.auth.decorators import login_required

@login_required
@csrf_protect
def data_view(request):
    full_name = f"{request.user.first_name} {request.user.last_name}"

    is_admin = EmployeeMaster.objects.filter(Admin=full_name).exists()
    is_partner = EmployeeMaster.objects.filter(Partner=full_name).exists()

    if not (is_admin or is_partner):
        return HttpResponseForbidden("You are not authorized to access this page.")
    
    # Default empty queryset
    employees = EmployeeMaster.objects.none()
    clients = ClientMaster.objects.none()

    is_team_leader = False  # flag
    # Role-based filtering (same as calendar_view)
    if EmployeeMaster.objects.filter(Admin=full_name).exists():
        employees = EmployeeMaster.objects.filter(Admin=full_name).distinct().order_by('Employee_name')
        clients = ClientMaster.objects.all().distinct().order_by('Client_name')
    elif EmployeeMaster.objects.filter(Partner=full_name).exists():
        employees = EmployeeMaster.objects.filter(Partner=full_name).distinct().order_by('Employee_name')
        clients = ClientMaster.objects.filter(Head=full_name).distinct().order_by('Client_name')
    elif EmployeeMaster.objects.filter(audit_manager=full_name).exists():
        employees = EmployeeMaster.objects.filter(audit_manager=full_name).distinct().order_by('Employee_name')
        clients = ClientMaster.objects.filter(audit_manager=full_name).distinct().order_by('Client_name')  # Team leader's clients only
        is_team_leader = True

    # Fetch all users and planning entries (no filter needed)
    employee_names = employees.values_list('Employee_name', flat=True)

    # Build Q objects for filtering users
    user_filter = Q()
    for emp_name in employee_names:
        parts = emp_name.strip().split(" ", 1)  # split into first and last name
        if len(parts) == 2:
            first, last = parts
        else:
            first, last = parts[0], ""  # only first name, no last name
        user_filter |= Q(first_name__iexact=first.strip(), last_name__iexact=last.strip())

    users = User.objects.filter(user_filter)


    context = {
        'employees': employees,
        'clients': clients,
        'users': users,
        'is_team_leader': is_team_leader,   # pass the flag
    }
    return render(request, 'data.html', context)



from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import EmployeeMaster, ClientMaster
from django.contrib.auth.models import User

@login_required
@csrf_protect
def save_employee(request):
    if request.method == "POST":
        # print("save_employee POST data:", request.POST)  # Add this
        data = request.POST
        obj = EmployeeMaster.objects.create(
            Employee_name=data.get('employee_name', ''),
            audit_manager=data.get('audit_manager', ''),
            Partner=data.get('partner', ''),
            Admin=data.get('admin', '')
        )
        return JsonResponse({'id': obj.pk})


@login_required
@csrf_protect
def update_employee(request):
    if request.method == "POST":
        # print("update_employee POST data:", request.POST)  # Add this
        obj = EmployeeMaster.objects.get(pk=request.POST.get('id'))
        obj.Employee_name = request.POST.get('employee_name', obj.Employee_name)
        obj.audit_manager = request.POST.get('audit_manager', obj.audit_manager)
        obj.Partner = request.POST.get('partner', obj.Partner)
        obj.Admin = request.POST.get('admin', obj.Admin)
        obj.save()
        return JsonResponse({'id': obj.pk})


@login_required
@csrf_protect
def delete_employee(request):
    if request.method == "POST":
        data = request.POST
        EmployeeMaster.objects.filter(pk=data.get('id')).delete()
        return JsonResponse({'deleted': True})

@login_required
@csrf_protect
def save_client(request):
    if request.method == "POST":
        data = request.POST
        obj = ClientMaster.objects.create(
            Client_name=data.get('client_name', ''),
            audit_manager=data.get('audit_manager', ''),
            Partner=data.get('partner', '')  # updated from 'head'
        )
        return JsonResponse({'id': obj.pk})

@login_required
@csrf_protect
def update_client(request):
    if request.method == "POST":
        data = request.POST
        obj = ClientMaster.objects.get(pk=data.get('id'))
        obj.Client_name = data.get('client_name', obj.Client_name)
        obj.audit_manager = data.get('audit_manager', obj.audit_manager)
        obj.Partner = data.get('partner', obj.Partner)  # updated from Head
        obj.save()
        return JsonResponse({'id': obj.pk})

@login_required
@csrf_protect
def delete_client(request):
    if request.method == "POST":
        data = request.POST
        ClientMaster.objects.filter(pk=data.get('id')).delete()
        return JsonResponse({'deleted': True})

@login_required
@csrf_protect
def save_user(request):  
    if request.method == "POST":
        data = request.POST
        # Create user with default password - change as needed
        obj = User.objects.create_user(
            username=data.get('username', ''),
            first_name=data.get('first_name', ''),
            last_name=data.get('last_name', ''),
            password='password123'
        )
        obj.is_superuser = data.get('is_superuser', 'No') == 'Yes'
        obj.is_staff = data.get('is_staff', 'No') == 'Yes'
        obj.save()
        return JsonResponse({'id': obj.pk})

@login_required
@csrf_protect
def update_user(request):
    if request.method == "POST":
        data = request.POST
        obj = User.objects.get(pk=data.get('id'))
        obj.username = data.get('username', obj.username)
        obj.first_name = data.get('first_name', obj.first_name)
        obj.last_name = data.get('last_name', obj.last_name)
        obj.is_superuser = data.get('is_superuser', 'No') == 'Yes'
        obj.is_staff = data.get('is_staff', 'No') == 'Yes'
        obj.save()
        return JsonResponse({'id': obj.pk})

@login_required
@csrf_protect
def delete_user(request):
    if request.method == "POST":
        data = request.POST
        User.objects.filter(pk=data.get('id')).delete()
        return JsonResponse({'deleted': True})


from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.shortcuts import render, redirect
from django.contrib import messages

@login_required
@csrf_protect
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()  # Save the new password
            update_session_auth_hash(request, user)  # Keep user logged in after password change
            messages.success(request, 'Your password was successfully updated!')
            return redirect('change_password')  # Or redirect wherever you want
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PasswordChangeForm(user=request.user)
    
    return render(request, 'change_password.html', {'form': form})

# views.py
from datetime import timedelta
from django.contrib import messages
from django.shortcuts import render, redirect
from .forms import HolidayForm
from .models import EmployeeMaster, Planning
from django.http import HttpResponseForbidden

@login_required
@csrf_protect
def add_holiday(request):
    full_name = f"{request.user.first_name} {request.user.last_name}"
    
    # Block non-admins
    if not EmployeeMaster.objects.filter(Admin=full_name).exists():
        return HttpResponseForbidden("You are not authorized to access this page.")
    
    if request.method == 'POST':
        form = HolidayForm(request.POST)
        if form.is_valid():
            from_date = form.cleaned_data['from_date']
            to_date = form.cleaned_data['to_date']
            fullday_halfday = form.cleaned_data['fullday_halfday']
            description = form.cleaned_data['description']
            action = request.POST.get('action')

            if action == 'add':
                employees = EmployeeMaster.objects.values_list('Employee_name', flat=True).distinct()
                current_date = from_date
                count = 0

                while current_date <= to_date:
                    for emp in employees:
                        Planning.objects.create(
                            date=current_date,
                            employee_name=emp,
                            client_name='Holiday',
                            fullday_halfday=fullday_halfday,
                            description=description
                        )
                        count += 1
                    current_date += timedelta(days=1)

                messages.success(request, f"Holiday added for {count} entries.")

            elif action == 'delete':
                deleted_count, _ = Planning.objects.filter(
                    date__range=(from_date, to_date),
                    client_name='Holiday'
                ).delete()
                messages.success(request, f"Deleted {deleted_count} holiday entries.")

            else:
                messages.error(request, "No valid action specified.")

            return redirect('add_holiday')

        else:
            # form invalid
            messages.error(request, "Please fix the errors below.")
    else:
        form = HolidayForm()

    return render(request, 'add_holiday.html', {'form': form})

#
# def calendar_view(request):
#     employees = EmployeeMaster.objects.values('Employee_name').distinct().order_by('Employee_name')
#     clients = ClientMaster.objects.values('Client_name').distinct().order_by('Client_name')
#     return render(request, 'calendar.html', {
#         'employees': employees,
#         'clients': clients,
#     })

from django.contrib.auth.decorators import login_required
from django.db.models import Q
from .models import EmployeeMaster, ClientMaster

@login_required
@csrf_protect
def calendar_view(request):
    full_name = f"{request.user.first_name} {request.user.last_name}"

    # Check role in priority order
    if EmployeeMaster.objects.filter(Admin=full_name).exists():
        employees = EmployeeMaster.objects.filter(Admin=full_name).values('Employee_name').distinct().order_by('Employee_name')
        clients = ClientMaster.objects.values('Client_name').distinct().order_by('Client_name')
    elif EmployeeMaster.objects.filter(Partner=full_name).exists():
        employees = EmployeeMaster.objects.filter(Partner=full_name).values('Employee_name').distinct().order_by('Employee_name')
        clients = ClientMaster.objects.filter(Head=full_name).values('Client_name').distinct().order_by('Client_name')
    elif EmployeeMaster.objects.filter(audit_manager=full_name).exists():
        employees = EmployeeMaster.objects.filter(audit_manager=full_name).values('Employee_name').distinct().order_by('Employee_name')
        clients = ClientMaster.objects.filter(audit_manager=full_name).values('Client_name').distinct().order_by('Client_name')
    else:
        employees = []  # No access
        clients = []
    # clients = ClientMaster.objects.values('Client_name').distinct().order_by('Client_name')

    return render(request, 'calendar.html', {
        'employees': employees,
        'clients': clients,
    })


from django.http import JsonResponse
from .models import Planning

@login_required
@csrf_protect
def get_planning_events(request):
    employee_name = request.GET.get('employee_name')
    month = int(request.GET.get('month'))
    year = int(request.GET.get('year'))

    plans = Planning.objects.filter(
        employee_name=employee_name,
        date__month=month,
        date__year=year
    )

    events = []
    
    for plan in plans:
        title_suffix = "- F" if "full day" in plan.fullday_halfday.lower() else "- H"
        title = f"{plan.client_name} {title_suffix}"
        client = plan.client_name.lower()
        if "leave" in client:
            color = "orange"
        elif "holiday" in client:
            color = "red"
        else:
            color = "#3788d8"  # FullCalendar default blue

        events.append({
            "id": plan.id,
            "title": title,
            "start": plan.date.isoformat(),
            "allDay": True,
            "color": color,
            "description": plan.description or "",  # <-- add description safely
            "clients": [plan.client_name],
        })

    return JsonResponse(events, safe=False)


from django.http import JsonResponse
from datetime import date, timedelta
from calendar import monthrange
from .models import Planning

@login_required
@csrf_protect
def get_planning_summary(request):
    employee_name = request.GET.get('employee_name')
    month = int(request.GET.get('month'))
    year = int(request.GET.get('year'))

    # Calculate total days in month
    num_days_in_month = monthrange(year, month)[1]

    # Calculate number of working days (Mon-Fri)
    working_days = num_days_in_month  # total days including Sundays
    sundays = []

    for day in range(1, num_days_in_month + 1):
        d = date(year, month, day)
        if d.weekday() == 6:  # Sunday
            sundays.append(d)


    # Initialize counters as floats to accumulate half days
    allocated = 0.0
    holiday = 0.0
    leave = 0.0
    pending = 0.0
    available = 0.0
    # Get all plans for employee for the month
    plans = Planning.objects.filter(
        employee_name=employee_name,
        date__year=year,
        date__month=month
    )

    # We'll track which Sundays have plans
    sunday_plans_dates = set(plan.date for plan in plans if plan.date.weekday() == 6)

    # Process each plan
    for plan in plans:
        client_lower = (plan.client_name or '').lower()
        day_value = 1.0 if plan.fullday_halfday and plan.fullday_halfday.lower() == 'full day' else 0.5

        if plan.date.weekday() == 6:  # Sunday
            # Plan on Sunday means allocated
            allocated += day_value
        else:
            # Non-Sunday days
            if 'holiday' in client_lower:
                holiday += day_value
            elif 'leave' in client_lower:
                leave += day_value
            else:
                allocated += day_value

    # Sundays without any plan count as holiday full day (1 each)
    sundays_without_plans = set(sundays) - sunday_plans_dates
    holiday += len(sundays_without_plans) * 1.0

    # Available = working days - allocated - holiday - leave
    # Note: working_days excludes Sundays already
    available = working_days - (holiday + leave)
    pending = available - allocated
    if available < 0:
        available = 0.0

    # Pending is 0 (or add your logic if available)
    

    return JsonResponse({
        'working_days': working_days,
        'available': round(available, 2),
        'allocated': round(allocated, 2),
        'pending': round(pending, 2),
        'holiday': round(holiday, 2),
        'leave': round(leave, 2),
    })




# import json
# from django.http import JsonResponse
# from django.views.decorators.csrf import csrf_exempt
# from .models import Planning
# from datetime import datetime,timedelta

# @csrf_exempt
# def save_planning(request):
#     if request.method == 'POST':
#         try:
#             data = json.loads(request.body)
#             event_id = data.get('event_id')  # <-- Get event_id here (optional)
#             employee_name = data.get('employee_name')
#             from_date = data.get('from_date')
#             to_date = data.get('to_date')
#             day_type = data.get('day_type')
#             clients = data.get('clients', [])
#             description = data.get('description', '')

#             from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
#             to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()

#             if event_id:
#                 # Update existing event (assuming single date and client)
#                 planning = Planning.objects.get(id=event_id)
#                 planning.employee_name = employee_name
#                 planning.date = from_date_obj  # or to_date_obj if date range not supported here
#                 planning.fullday_halfday = day_type
#                 planning.client_name = clients[0] if clients else planning.client_name
#                 planning.description = description
#                 planning.save()
#             else:
#                 # Create new events for each date/client in range
#                 delta = (to_date_obj - from_date_obj).days
#                 for i in range(delta + 1):
#                     current_date = from_date_obj + timedelta(days=i)
#                     for client in clients:
#                         Planning.objects.update_or_create(
#                             date=current_date,
#                             employee_name=employee_name,
#                             client_name=client,
#                             defaults={'fullday_halfday': day_type, 'description': description},
#                         )

#             return JsonResponse({'message': 'Planning saved successfully.'})
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=400)

#     return JsonResponse({'error': 'Invalid method'}, status=405)




from django.db.models import Q
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Planning
from datetime import datetime,timedelta

@login_required
@csrf_protect
def save_planning(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            event_id = data.get('event_id')
            employee_name = data.get('employee_name')
            from_date = data.get('from_date')
            to_date = data.get('to_date')
            day_type = data.get('day_type')
            clients = data.get('clients', [])
            description = data.get('description', '')

            from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
            to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()

            # Generate all dates in range
            if from_date_obj == to_date_obj:
                date_list = [from_date_obj]  # 🔸 Only one entry for the day
            else:
                date_list = [from_date_obj + timedelta(days=i) for i in range((to_date_obj - from_date_obj).days + 1)]
                # 🔸 Adds entries for each day in range


            # Check existing plannings for these dates and employee
            existing_plans = Planning.objects.filter(
                employee_name=employee_name,
                date__in=date_list,
            )

            # Helper counts
            full_day_count = existing_plans.filter(fullday_halfday__iexact='full day').count()
            half_day_count = existing_plans.filter(fullday_halfday__iexact='half day').count()

            # Also check if any full day exists for these dates for this employee
            full_day_exists = full_day_count > 0

            # Business rules check
            # Business rules check per date
            if day_type == 'full day':
                for d in date_list:
                    existing_full_day = existing_plans.filter(date=d, fullday_halfday__iexact='full day').exists()
                    existing_half_day_count = existing_plans.filter(date=d, fullday_halfday__iexact='half day').count()
                    if existing_full_day or existing_half_day_count > 0:
                        return JsonResponse({'error': f'Conflict on {d}: Full day planning already exists.'}, status=400)
                if len(clients) > 1:
                    return JsonResponse({'error': 'Only one client allowed for full day planning.'}, status=400)

            elif day_type == 'half day':
                for d in date_list:
                    existing_full_day = existing_plans.filter(date=d, fullday_halfday__iexact='full day').exists()
                    existing_half_day_count = existing_plans.filter(date=d, fullday_halfday__iexact='half day').count()


                    new_plans_for_date = len(clients)

                    if existing_full_day:
                        return JsonResponse({'error': f'Full day planning exists on {d}. Cannot add half day.'}, status=400)
                    if existing_half_day_count + new_plans_for_date > 2:
                        return JsonResponse({'error': f'Maximum two half day plannings allowed on {d}. Cannot add more.'}, status=400)


            # Now proceed with save/update

            if event_id:
                # Assuming event_id corresponds to multiple Planning records? 
                # Since you store event_id on frontend, backend should map to actual Planning records.
                # If you store event_id in your model, use it; else delete by employee+date+clients

                # Delete existing plannings that fall in this date range, employee, and clients
                Planning.objects.filter(
                    employee_name=employee_name,
                    date__range=(from_date_obj, to_date_obj),
                    client_name__in=clients
                ).delete()
            day_type = data.get('day_type')


            # Create new Planning entries for each date and client
            for current_date in date_list:
                for client in clients:
                    Planning.objects.create(
                        employee_name=employee_name,
                        date=current_date,
                        fullday_halfday=day_type,
                        client_name=client,
                        description=description
                    )

            return JsonResponse({'message': 'Planning saved successfully.'})

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Invalid method'}, status=405)




@login_required
@csrf_protect
def delete_planning(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            employee_name = data.get('employee_name')
            from_date = data.get('from_date')
            to_date = data.get('to_date')
            clients = data.get('clients', [])  # ⬅️ use this list

            from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
            to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()

            # Delete all matching records in range
            Planning.objects.filter(
                employee_name=employee_name,
                date__range=(from_date_obj, to_date_obj),
                client_name__in=clients
            ).delete()

            return JsonResponse({'message': 'Planning deleted successfully.'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Invalid method'}, status=405)








import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment
from django.http import HttpResponse
from django.db import connection
from datetime import date, timedelta
import calendar
import os
from django.conf import settings

@login_required
@csrf_protect
def export_planning_excel(request):
    employee_name = request.GET.get("employee_name")
    month = request.GET.get("month")
    year = request.GET.get("year")

    if not employee_name or not month or not year:
        return HttpResponse("Missing parameters", status=400)

    year = int(year)
    month = int(month)

    # Date range for selected month
    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])
    all_dates = [first_day + timedelta(days=i) for i in range((last_day - first_day).days + 1)]

    # Define holidays (example)
    holidays = [date(year, month, 15)]  # Modify or load dynamically

    # Query planning data
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT date, client_name, fullday_halfday, description
            FROM planning
            WHERE employee_name = %s 
            AND EXTRACT(YEAR FROM date) = %s 
            AND EXTRACT(MONTH FROM date) = %s

        """, [employee_name, year, month])
        rows = cursor.fetchall()

    # Process data
    data_by_date = {}
    for row in rows:
        d, client, day_type, desc = row
        suffix = " - F" if (day_type and day_type.lower() == "full day") else " - H" if day_type else ""
        entry = {
            "client": f"{client}{suffix}" if client else "",
            "day_type": day_type,
            "desc": desc
        }
        if d in data_by_date:
            data_by_date[d].append(entry)
        else:
            data_by_date[d] = [entry]


    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planning Report"

    # Add image in A1
    img_path = os.path.join(settings.BASE_DIR, "dashboard", "static", "projectimage", "rspc.png")
    if os.path.exists(img_path):
        try:
            img = XLImage(img_path)
            img.width = 250
            img.height = 43
            ws.add_image(img, "A1")
            ws.row_dimensions[1].height = 45
            ws.column_dimensions['A'].width = 55
        except Exception as e:
            print("Error adding image:", e)
    else:
        print("Image not found:", img_path)

        
    # Title in A2
    ws["A2"] = f"{employee_name} - Monthly Plan {calendar.month_name[month]} {year}"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Header at row 3
    headers = ["Date", "Day", "Client Name", "Description"]
    ws.append(headers)

    # Apply header style
    header_fill = PatternFill(start_color="000C66", end_color="000C66", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data colors
    red_fill = PatternFill(start_color="FFDDDD", end_color="FFDDDD", fill_type="solid")     # Light red
    holiday_red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")     # Light red
    orange_fill = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid")  # Orange

    # Fill data from row 4 onward
    for d in all_dates:
        formatted_date = d.strftime("%d-%m-%Y")
        day_name = d.strftime("%A")
        entries = data_by_date.get(d, [])

        if entries:
            # Combine clients and descriptions separated by newlines
            combined_clients = "\n".join(entry.get("client", "") for entry in entries)
            combined_descs = "\n".join(entry.get("desc", "") for entry in entries)
        else:
            combined_clients = ""
            combined_descs = ""

        ws.append([formatted_date, day_name, combined_clients, combined_descs])
        row_idx = ws.max_row

        # Apply row fill
        if day_name == "Sunday":
            fill = red_fill
        elif "leave" in combined_clients.lower():
            fill = orange_fill
        elif "holiday" in combined_clients.lower():
            fill = holiday_red_fill
        else:
            fill = None

        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

        # Optional: Enable wrap text on client and desc columns so multiline shows correctly
        ws.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True)
        ws.cell(row=row_idx, column=4).alignment = Alignment(wrap_text=True)


    # Column widths
    # for col_num in range(1, len(headers) + 1):
    #     col_letter = get_column_letter(col_num)
    #     max_length = 0
    #     for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=col_num, max_col=col_num):
    #         for cell in row:
    #             if cell.value:
    #                 max_length = max(max_length, len(str(cell.value)))
    #     # Add a little extra space
    #     adjusted_width = max_length + 5
    #     ws.column_dimensions[col_letter].width = adjusted_width

    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 12
    # Generate response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{employee_name}_Planning_{year}_{month:02d}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response





from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
from datetime import date, timedelta
import calendar
from django.http import HttpResponse
from django.db import connection
from reportlab.platypus import Image as RLImage
import os

@login_required
@csrf_protect
def export_planning_pdf(request):
    employee_name = request.GET.get("employee_name")
    month = request.GET.get("month")
    year = request.GET.get("year")

    if not employee_name or not month or not year:
        return HttpResponse("Missing parameters", status=400)

    year = int(year)
    month = int(month)

    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])
    all_dates = [first_day + timedelta(days=i) for i in range((last_day - first_day).days + 1)]

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT date, client_name, fullday_halfday, description
            FROM planning
            WHERE employee_name = %s 
            AND EXTRACT(YEAR FROM date) = %s 
            AND EXTRACT(MONTH FROM date) = %s

        """, [employee_name, year, month])
        rows = cursor.fetchall()

    data_by_date = {}
    for row in rows:
        d, client, day_type, desc = row
        suffix = " - F" if (day_type and day_type.lower() == "full day") else " - H" if day_type else ""
        entry = {
            "client": f"{client}{suffix}" if client else "",
            "desc": desc or ""
        }
        data_by_date.setdefault(d, []).append(entry)

    # Start PDF generation
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    # Load logo image if exists
    logo_path = os.path.join(settings.BASE_DIR, "dashboard", "static", "projectimage", "rspc.png")
    logo_width = 120
    logo_height = 40
    if os.path.exists(logo_path):
        logo = RLImage(logo_path, width=logo_width, height=logo_height)
    else:
        logo = Paragraph("", styles['Normal'])

    # Title paragraph
    title_text = f"<b>{employee_name} - Monthly Plan for {calendar.month_name[month]} {year}</b>"
    title = Paragraph(title_text, styles['Title'])

    # Table for header: logo left, title right
    header_table = Table([[logo, title]], colWidths=[logo_width + 10, 400])
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
    ]))

    elements.append(header_table)
    elements.append(Spacer(1, 12))

    # Table header for planning data
    table_data = [["Date", "Day", "Client Name(s)", "Description"]]

    for d in all_dates:
        formatted_date = d.strftime("%d-%m-%Y")
        day_name = d.strftime("%A")
        entries = data_by_date.get(d, [])
        combined_clients = "\n".join(entry["client"] for entry in entries) or "-"
        combined_descs = "\n".join(entry["desc"] for entry in entries) or "-"
        table_data.append([formatted_date, day_name, combined_clients, combined_descs])

    # Create data table
    table = Table(table_data, colWidths=[80, 80, 180, 180])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#000C66")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    # Highlight Sundays, leaves, holidays
    for i, row in enumerate(table_data[1:], start=1):
        day = row[1].lower()
        client_info = row[2].lower()
        if "sunday" in day:
            table.setStyle([("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FFDDDD"))])
        elif "leave" in client_info:
            table.setStyle([("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FFBF00"))])
        elif "holiday" in client_info:
            table.setStyle([("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FF0000"))])

    elements.append(table)

    doc.build(elements)

    buffer.seek(0)
    filename = f"{employee_name}_Planning_{year}_{month:02d}.pdf"
    return HttpResponse(buffer, content_type='application/pdf', headers={'Content-Disposition': f'attachment; filename="{filename}"'})






# def quarter_view(request):
#     employees = EmployeeMaster.objects.values('Employee_name').distinct().order_by('Employee_name')
#     clients = ClientMaster.objects.values('Client_name').distinct().order_by('Client_name')
#     return render(request, 'quarterview.html', {
#         'employees': employees,
#         'clients': clients,
#     })




from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import EmployeeMaster, ClientMaster


@login_required
@csrf_protect
def quarter_view(request):
    full_name = f"{request.user.first_name} {request.user.last_name}"

    # ================================
    # EMPLOYEE FILTERING BY ROLE
    # ================================
    if EmployeeMaster.objects.filter(Admin=full_name).exists():
        employees = EmployeeMaster.objects.filter(Admin=full_name).values('Employee_name').distinct().order_by('Employee_name')
        
        # ADMIN sees all clients
        clients = ClientMaster.objects.values('Client_name').distinct().order_by('Client_name')

    elif EmployeeMaster.objects.filter(Partner=full_name).exists():
        employees = EmployeeMaster.objects.filter(Partner=full_name).values('Employee_name').distinct().order_by('Employee_name')
        
        # PARTNER sees clients where head = full_name
        clients = ClientMaster.objects.filter(Head=full_name).values('Client_name').distinct().order_by('Client_name')

    elif EmployeeMaster.objects.filter(audit_manager=full_name).exists():
        employees = EmployeeMaster.objects.filter(audit_manager=full_name).values('Employee_name').distinct().order_by('Employee_name')
        
        # TEAM LEADER sees clients where audit_manager = full_name
        clients = ClientMaster.objects.filter(audit_manager=full_name).values('Client_name').distinct().order_by('Client_name')

    else:
        employees = []
        clients = []

    return render(request, 'quarterview.html', {
        'employees': employees,
        'clients': clients,
    })




from django.http import JsonResponse
from django.utils.dateparse import parse_date
from .models import Planning

@login_required
@csrf_protect
def get_planning_data(request):
    employee_name = request.GET.get('employee_name')
    from_date_str = request.GET.get('from_date')  # expected YYYY-MM-DD
    to_date_str = request.GET.get('to_date')      # expected YYYY-MM-DD

    if not (employee_name and from_date_str and to_date_str):
        return JsonResponse({"error": "Missing parameters"}, status=400)

    from_date = parse_date(from_date_str)
    to_date = parse_date(to_date_str)

    plans = Planning.objects.filter(
        employee_name=employee_name,
        date__range=(from_date, to_date)
    )

    # Structure data by date string (DD-MM-YYYY) => list of event titles
    data_by_date = {}

    for plan in plans:
        suffix = "- F" if "full day" in plan.fullday_halfday.lower() else "- H"
        title = f"{plan.client_name} {suffix}"
        date_key = plan.date.strftime("%d-%m-%Y")
        if date_key not in data_by_date:
            data_by_date[date_key] = []
        data_by_date[date_key].append(title)

    return JsonResponse(data_by_date)





from openpyxl import Workbook
from .models import Planning
from openpyxl.styles import Alignment,Font
from datetime import date, timedelta


@login_required
@csrf_protect
def quarterview_excel(request):
    

    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    employee_name = request.GET.get('employee_name')

    if not (from_date_str and to_date_str and employee_name):
        return HttpResponse("Missing parameters", status=400)

    from_date = datetime.strptime(from_date_str, "%Y-%m-%d").date()
    to_date = datetime.strptime(to_date_str, "%Y-%m-%d").date()
    # print(from_date)


    from_date_dd = from_date.strftime("%d-%m-%Y")
    to_date_dd = to_date.strftime("%d-%m-%Y")


    # Get list of 3 months
    months = []
    current = date(from_date.year, from_date.month, 1)
    for _ in range(3):
        months.append(current)
        if current.month == 12:
            current = date(current.year + 1, 1, 1)
        else:
            current = date(current.year, current.month + 1, 1)

    # Get planning data
    plans = Planning.objects.filter(
        employee_name=employee_name,
        date__range=(from_date, to_date)
    )

    # Structure data by DD-MM-YYYY => list of titles
    data_by_date = {}
    for plan in plans:
        suffix = "-F" if "full" in plan.fullday_halfday.lower() else "-H"
        title = f"{plan.client_name} {suffix}"
        date_key = plan.date.strftime("%d-%m-%Y")
        data_by_date.setdefault(date_key, []).append(title)

    # Excel setup
    wb = Workbook()
    ws = wb.active
    ws.title = f"Quarter Plan"

    # === Add image ===
    img_path = os.path.join(settings.BASE_DIR, "dashboard", "static", "projectimage", "rspc.png")
    # print(f"Image absolute path: {img_path}")
    # print("Image exists:", os.path.exists(img_path))

    if os.path.exists(img_path):
        img = XLImage(img_path)
        img.width = 250
        img.height = 43
        ws.add_image(img, "A1")
        ws.row_dimensions[1].height = 45
        ws.column_dimensions['A'].width = 55

    # === Add title in A2 ===

    ws["A2"] = f"{employee_name} - Quarter Plan For {from_date_dd} to {to_date_dd}"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # === Add headers at row 3 ===
    headers = ['Date'] + [m.strftime('%B') for m in months]
    ws.append(headers)

    header_fill = PatternFill(start_color="000C66", end_color="000C66", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Fill colors
    holiday_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    leave_fill = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid")
    sunday_fill = PatternFill(start_color="FFDDDD", end_color="FFDDDD", fill_type="solid")
    event_font = Font(name="Courier New", bold=True, size=11)

    # === Write data rows for each day ===
    for day in range(1, 32):
        row = [str(day)]  # First column is day number
        for month in months:
            try:
                cell_date = date(month.year, month.month, day)
            except ValueError:
                row.append('')
                continue

            date_key = cell_date.strftime("%d-%m-%Y")
            weekday = cell_date.strftime('%a')
            events = data_by_date.get(date_key, [])
            event_text = ', '.join(events) if events else ''
            cell_value = f"{weekday}       {event_text}".strip()
            row.append(cell_value)

        ws.append(row)

        for col_idx in range(2, 5):  # Columns B, C, D
            try:
                cell_date = date(months[col_idx - 2].year, months[col_idx - 2].month, day)
            except ValueError:
                continue

            cell = ws.cell(row=day + 3, column=col_idx)
            cell.font = event_font
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            if cell_date.weekday() == 6:  # Sunday
                cell.fill = sunday_fill

            val = cell.value.lower() if cell.value else ''
            if 'holiday' in val:
                cell.fill = holiday_fill
            elif 'leave' in val:
                cell.fill = leave_fill

    # Autosize columns
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 12

    # Final response

    output = BytesIO()
    wb.save(output)
    output.seek(0)  # Go back to the start of the BytesIO buffer

    # Create response with correct content type
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    filename = f"{employee_name}_Quarter_Plan_{from_date_dd}_to_{to_date_dd}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response






from io import BytesIO
from datetime import datetime, date
import os
from django.http import HttpResponse
from django.conf import settings
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
)
from reportlab.lib.styles import getSampleStyleSheet
from .models import Planning  # adjust import as needed


@login_required
@csrf_protect
def quarterview_pdf(request):
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    employee_name = request.GET.get('employee_name')

    if not (from_date_str and to_date_str and employee_name):
        return HttpResponse("Missing parameters", status=400)

    from_date = datetime.strptime(from_date_str, "%Y-%m-%d").date()
    to_date = datetime.strptime(to_date_str, "%Y-%m-%d").date()
    from_date_dd = from_date.strftime("%d-%m-%Y")
    to_date_dd = to_date.strftime("%d-%m-%Y")

    # Get 3 consecutive months
    months = []
    current = date(from_date.year, from_date.month, 1)
    for _ in range(3):
        months.append(current)
        if current.month == 12:
            current = date(current.year + 1, 1, 1)
        else:
            current = date(current.year, current.month + 1, 1)

    plans = Planning.objects.filter(
        employee_name=employee_name,
        date__range=(from_date, to_date)
    )

    data_by_date = {}
    for plan in plans:
        suffix = "-F" if "full" in plan.fullday_halfday.lower() else "-H"
        title = f"{plan.client_name} {suffix}"
        date_key = plan.date.strftime("%d-%m-%Y")
        data_by_date.setdefault(date_key, []).append(title)

    # === PDF Generation ===
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    # Logo + Title Header
    logo_path = os.path.join(settings.BASE_DIR, "dashboard", "static", "projectimage", "rspc.png")
    logo_width, logo_height = 120, 40
    if os.path.exists(logo_path):
        logo = RLImage(logo_path, width=logo_width, height=logo_height)
    else:
        logo = Paragraph("", styles['Normal'])

    title_text = f"<b>{employee_name} - Quarter Plan ({from_date_dd} to {to_date_dd})</b>"
    title = Paragraph(title_text, styles['Title'])
    header_table = Table([[logo, title]], colWidths=[logo_width + 10, 400])
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
    ]))

    elements.append(header_table)
    elements.append(Spacer(1, 12))

    # Table Headers
    header = ['Date'] + [m.strftime('%B') for m in months]
    table_data = [header]

    for day in range(1, 32):
        row = [str(day)]
        for month in months:
            try:
                cell_date = date(month.year, month.month, day)
            except ValueError:
                row.append('')
                continue

            date_key = cell_date.strftime("%d-%m-%Y")
            weekday = cell_date.strftime('%a')
            events = data_by_date.get(date_key, [])
            event_text = ', '.join(events)
            cell_value = f"{weekday}      {event_text}" if event_text else weekday
            row.append(cell_value)

        table_data.append(row)

    table = Table(table_data, repeatRows=1, colWidths=[50, 120, 120, 120])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#000C66")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Courier"),  # 👈 use monospaced font for data rows
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))


    # Highlight Sundays, Holidays, Leaves
    for r in range(1, len(table_data)):
        for c in range(1, len(months) + 1):
            cell = table_data[r][c].lower()
            if 'holiday' in cell:
                table.setStyle([("BACKGROUND", (c, r), (c, r), colors.HexColor("#FF0000"))])
            elif 'leave' in cell:
                table.setStyle([("BACKGROUND", (c, r), (c, r), colors.HexColor("#FFBF00"))])
            elif cell.startswith("sun"):
                table.setStyle([("BACKGROUND", (c, r), (c, r), colors.HexColor("#FFDDDD"))])





    elements.append(table)
    doc.build(elements)

    filename = f"{employee_name}_Quarter_Plan_{from_date_dd}_to_{to_date_dd}.pdf"
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf', headers={
        'Content-Disposition': f'attachment; filename="{filename}"'
    })





from django.shortcuts import render
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import EmployeeMaster, ClientMaster, Planning
from collections import defaultdict
import pandas as pd
from datetime import datetime

@login_required
@csrf_protect
def allocated_resources_view(request):
    full_name = f"{request.user.first_name} {request.user.last_name}"
    is_admin = EmployeeMaster.objects.filter(Admin=full_name).exists()
    is_partner = EmployeeMaster.objects.filter(Partner=full_name).exists()

    if not (is_admin or is_partner):
        return HttpResponseForbidden("You are not authorized to access this page.")
    
    return render(request, 'allocated_resources.html')



from collections import defaultdict
from datetime import datetime, timedelta
import calendar
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import EmployeeMaster, ClientMaster

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from collections import defaultdict
from datetime import datetime, timedelta
import calendar
from .models import EmployeeMaster, ClientMaster

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from datetime import datetime, timedelta
import calendar

from .models import EmployeeMaster, ClientMaster, Planning

@login_required
@csrf_protect
def allocated_data(request):
    full_name = f"{request.user.first_name} {request.user.last_name}"

    is_admin = EmployeeMaster.objects.filter(Admin=full_name).exists()
    is_partner = EmployeeMaster.objects.filter(Partner=full_name).exists()

    if not (is_admin or is_partner):
        return HttpResponseForbidden("You are not authorized to access this page.")
    
    is_team_leader = False

    # Your existing role-based employee & client fetching
    if EmployeeMaster.objects.filter(Admin=full_name).exists():
        employees = EmployeeMaster.objects.filter(Admin=full_name).values_list('Employee_name', flat=True).distinct()
        clients = ClientMaster.objects.values_list('Client_name', flat=True).distinct()
    elif EmployeeMaster.objects.filter(Partner=full_name).exists():
        employees = EmployeeMaster.objects.filter(Partner=full_name).values_list('Employee_name', flat=True).distinct()
        clients = ClientMaster.objects.filter(Head=full_name).values_list('Client_name', flat=True).distinct()
    elif EmployeeMaster.objects.filter(audit_manager=full_name).exists():
        employees = EmployeeMaster.objects.filter(audit_manager=full_name).values_list('Employee_name', flat=True).distinct()
        clients = ClientMaster.objects.filter(audit_manager=full_name).values_list('Client_name', flat=True).distinct()
        is_team_leader = True
    else:
        employees = []
        clients = []

    emp_names = list(employees)
    client_names = list(clients)
    client_names.extend(["Holiday", "Leave"])

    from_date = datetime.strptime(request.GET.get('fromDate'), '%Y-%m-%d')
    to_date = datetime.strptime(request.GET.get('toDate'), '%Y-%m-%d')
    date_range = [from_date + timedelta(days=i) for i in range((to_date - from_date).days + 1)]

    # Get all distinct client-employee pairs in planning data in the range
    planning_data = Planning.objects.filter(
        employee_name__in=emp_names,
        client_name__in=client_names,
        date__range=(from_date, to_date)
    ).values('date', 'employee_name', 'client_name')

    # Extract unique client-employee pairs for headers
    client_employee_pairs = set()
    for entry in planning_data:
        client_employee_pairs.add((entry['client_name'], entry['employee_name']))

    # Sort pairs for consistent header order
    client_employee_pairs = sorted(client_employee_pairs, key=lambda x: (x[0], x[1]))

    # Build a list of header keys and display names
    # Key format: "client_employee" to be unique in JS
    columns = ['date', 'day']
    col_display = ['date', 'day']
    for client, emp in client_employee_pairs:
        key = f"{client}_{emp.replace(' ', '_')}"  # e.g. ICICI_Pankti_Shah
        columns.append(key)
        col_display.append(client)  # Show only client name in header

    # Build lookup for quick filling
    # plan_lookup[date_str][(client, employee)] = True
    plan_lookup = {}
    for entry in planning_data:
        date_str = entry['date'].strftime('%Y-%m-%d')
        key = (entry['client_name'], entry['employee_name'])
        plan_lookup.setdefault(date_str, {})[key] = True

    # Prepare rows
    rows = []
    for date in date_range:
        date_str = date.strftime('%Y-%m-%d')
        row = {
            'date': date_str,
            'day': calendar.day_name[date.weekday()],
        }
        for client, emp in client_employee_pairs:
            key = f"{client}_{emp.replace(' ', '_')}"
            if plan_lookup.get(date_str, {}).get((client, emp), False):
                row[key] = emp  # or 'X' or employee name
            else:
                row[key] = ""
        rows.append(row)

    return JsonResponse({
        'columns': columns,
        'col_display': col_display,  # For frontend header text (only client names)
        'rows': rows,
        'is_team_leader': is_team_leader,
    })





from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime, timedelta
import calendar
import os
from django.conf import settings
from .models import EmployeeMaster, ClientMaster, Planning


@login_required
@csrf_protect
def export_allocated_data_excel(request):
    full_name = f"{request.user.first_name} {request.user.last_name}"

   
    is_admin = EmployeeMaster.objects.filter(Admin=full_name).exists()
    is_partner = EmployeeMaster.objects.filter(Partner=full_name).exists()

    if not (is_admin or is_partner):
        return HttpResponseForbidden("You are not authorized to access this page.")
    
    # Role-based fetching
    if EmployeeMaster.objects.filter(Admin=full_name).exists():
        employees = EmployeeMaster.objects.filter(Admin=full_name).values_list('Employee_name', flat=True).distinct()
        clients = ClientMaster.objects.values_list('Client_name', flat=True).distinct()
    elif EmployeeMaster.objects.filter(Partner=full_name).exists():
        employees = EmployeeMaster.objects.filter(Partner=full_name).values_list('Employee_name', flat=True).distinct()
        clients = ClientMaster.objects.filter(Head=full_name).values_list('Client_name', flat=True).distinct()
    elif EmployeeMaster.objects.filter(audit_manager=full_name).exists():
        employees = EmployeeMaster.objects.filter(audit_manager=full_name).values_list('Employee_name', flat=True).distinct()
        clients = ClientMaster.objects.filter(audit_manager=full_name).values_list('Client_name', flat=True).distinct()
    else:
        employees = []
        clients = []

    emp_names = list(employees)
    client_names = list(clients)
    client_names.extend(["Holiday", "Leave"])

    from_date = datetime.strptime(request.GET.get('fromDate'), '%Y-%m-%d')
    to_date = datetime.strptime(request.GET.get('toDate'), '%Y-%m-%d')
    date_range = [from_date + timedelta(days=i) for i in range((to_date - from_date).days + 1)]

    planning_data = Planning.objects.filter(
        employee_name__in=emp_names,
        client_name__in=client_names,
        date__range=(from_date, to_date)
    ).values('date', 'employee_name', 'client_name')

    client_employee_pairs = sorted(
        {(entry['client_name'], entry['employee_name']) for entry in planning_data},
        key=lambda x: (x[0], x[1])
    )

    columns = ['Date', 'Day']
    for client, emp in client_employee_pairs:
        columns.append(client)

    plan_lookup = {}
    for entry in planning_data:
        date_str = entry['date'].strftime('%Y-%m-%d')
        key = (entry['client_name'], entry['employee_name'])
        plan_lookup.setdefault(date_str, {})[key] = True

    # Workbook and Sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "AllocatedData"

    # === IMAGE in A1 ===
    img_path = os.path.join(settings.BASE_DIR, "dashboard", "static", "projectimage", "rspc.png")
    if os.path.exists(img_path):
        img = XLImage(img_path)
        img.width = 250
        img.height = 43
        ws.add_image(img, "A1")
        ws.row_dimensions[1].height = 45
        ws.column_dimensions['A'].width = 55

    # === TITLE in A2 ===
    from_date_dd = from_date.strftime('%d-%m-%Y')
    to_date_dd = to_date.strftime('%d-%m-%Y')
    ws["A2"] = f"Allocated Resources From {from_date_dd} To {to_date_dd}"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    header_fill = PatternFill(start_color="000C66", end_color="000C66", fill_type="solid")

    # === HEADER in row 3 ===
    for col_num, value in enumerate(columns, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = value
        cell.font = Font(bold=True,color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill

    # === DATA starting from row 4 ===
    for idx, date in enumerate(date_range, start=4):
        date_str = date.strftime('%Y-%m-%d')
        row = [
            date,
            calendar.day_abbr[date.weekday()]  # Short day name
        ]
        for client, emp in client_employee_pairs:
            if plan_lookup.get(date_str, {}).get((client, emp), False):
                row.append(emp)
            else:
                row.append("")

        
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=idx, column=col_num)
            cell.value = value
            if col_num == 1:  # Date column formatting
                cell.number_format = 'DD-MM-YYYY'
                cell.alignment = Alignment(horizontal='center')

        if calendar.day_abbr[date.weekday()] == "Sun":
            fill = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")  # Coral Red
            for col_num in range(1, len(row) + 1):
                ws.cell(row=idx, column=col_num).fill = fill

    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 12

    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter  # e.g. 'A', 'B', etc.
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        # Adjust column width (+2 for padding)
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    # === Response as Excel File ===
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Allocated Resources From {from_date_dd} To {to_date_dd}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response








@login_required
@csrf_protect
def unallocated_resources_view(request):
    full_name = f"{request.user.first_name} {request.user.last_name}"
    is_admin = EmployeeMaster.objects.filter(Admin=full_name).exists()
    is_partner = EmployeeMaster.objects.filter(Partner=full_name).exists()

    if not (is_admin or is_partner):
        return HttpResponseForbidden("You are not authorized to access this page.")
    
    return render(request, 'unallocated_resources.html')




@login_required
@csrf_protect
def unallocated_data(request):
    full_name = f"{request.user.first_name} {request.user.last_name}"


    is_admin = EmployeeMaster.objects.filter(Admin=full_name).exists()
    is_partner = EmployeeMaster.objects.filter(Partner=full_name).exists()

    if not (is_admin or is_partner):
        return HttpResponseForbidden("You are not authorized to access this page.")
    
    is_team_leader =False

    # Your existing role-based employee & client fetching
    if EmployeeMaster.objects.filter(Admin=full_name).exists():
        employees = EmployeeMaster.objects.filter(Admin=full_name).values_list('Employee_name', flat=True).distinct()
        clients = ClientMaster.objects.values_list('Client_name', flat=True).distinct()
    elif EmployeeMaster.objects.filter(Partner=full_name).exists():
        employees = EmployeeMaster.objects.filter(Partner=full_name).values_list('Employee_name', flat=True).distinct()
        clients = ClientMaster.objects.filter(Head=full_name).values_list('Client_name', flat=True).distinct()
    elif EmployeeMaster.objects.filter(audit_manager=full_name).exists():
        employees = EmployeeMaster.objects.filter(audit_manager=full_name).values_list('Employee_name', flat=True).distinct()
        clients = ClientMaster.objects.filter(audit_manager=full_name).values_list('Client_name', flat=True).distinct()

    else:
        employees = []
        clients = []

    emp_names = list(employees)
    client_names = list(clients)
    client_names.extend(["Holiday", "Leave"])

    from_date = datetime.strptime(request.GET.get('fromDate'), '%Y-%m-%d')
    to_date = datetime.strptime(request.GET.get('toDate'), '%Y-%m-%d')
    date_range = [from_date + timedelta(days=i) for i in range((to_date - from_date).days + 1)]

    # Get all distinct client-employee pairs in planning data in the range
    planning_data = Planning.objects.filter(
        employee_name__in=emp_names,
        client_name__in=client_names,
        date__range=(from_date, to_date)
    ).values('date', 'employee_name', 'client_name')

    # Extract unique client-employee pairs for headers
    client_employee_pairs = set()
    for entry in planning_data:
        client_employee_pairs.add((entry['client_name'], entry['employee_name']))

    # Sort pairs for consistent header order
    client_employee_pairs = sorted(client_employee_pairs, key=lambda x: (x[0], x[1]))

    # Build a list of header keys and display names
    # Key format: "client_employee" to be unique in JS
    columns = ['date', 'day']
    col_display = ['date', 'day']
    for client, emp in client_employee_pairs:
        key = f"{client}_{emp.replace(' ', '_')}"  # e.g. ICICI_Pankti_Shah
        columns.append(key)
        col_display.append(client)  # Show only client name in header

    # Build lookup for quick filling
    # plan_lookup[date_str][(client, employee)] = True
    plan_lookup = {}
    for entry in planning_data:
        date_str = entry['date'].strftime('%Y-%m-%d')
        key = (entry['client_name'], entry['employee_name'])
        plan_lookup.setdefault(date_str, {})[key] = True

    # Prepare rows
    rows = []
    for date in date_range:
        date_str = date.strftime('%Y-%m-%d')
        row = {
            'date': date_str,
            'day': calendar.day_name[date.weekday()],
        }
        for client, emp in client_employee_pairs:
            key = f"{client}_{emp.replace(' ', '_')}"
            if not plan_lookup.get(date_str, {}).get((client, emp), False):
                row[key] = emp  # or 'X' or employee name
            else:
                row[key] = ""
        rows.append(row)

    return JsonResponse({
        'columns': columns,
        'col_display': col_display,  # For frontend header text (only client names)
        'rows': rows,
        'is_team_leader': is_team_leader,
    })




@login_required
@csrf_protect
def export_unallocated_data_excel(request):
    full_name = f"{request.user.first_name} {request.user.last_name}"

    is_admin = EmployeeMaster.objects.filter(Admin=full_name).exists()
    is_partner = EmployeeMaster.objects.filter(Partner=full_name).exists()

    if not (is_admin or is_partner):
        return HttpResponseForbidden("You are not authorized to access this page.")

    # Role-based fetching
    if EmployeeMaster.objects.filter(Admin=full_name).exists():
        employees = EmployeeMaster.objects.filter(Admin=full_name).values_list('Employee_name', flat=True).distinct()
        clients = ClientMaster.objects.values_list('Client_name', flat=True).distinct()
    elif EmployeeMaster.objects.filter(Partner=full_name).exists():
        employees = EmployeeMaster.objects.filter(Partner=full_name).values_list('Employee_name', flat=True).distinct()
        clients = ClientMaster.objects.filter(Head=full_name).values_list('Client_name', flat=True).distinct()
    elif EmployeeMaster.objects.filter(audit_manager=full_name).exists():
        employees = EmployeeMaster.objects.filter(audit_manager=full_name).values_list('Employee_name', flat=True).distinct()
        clients = ClientMaster.objects.filter(audit_manager=full_name).values_list('Client_name', flat=True).distinct()
    else:
        employees = []
        clients = []

    emp_names = list(employees)
    client_names = list(clients)
    client_names.extend(["Holiday", "Leave"])

    from_date = datetime.strptime(request.GET.get('fromDate'), '%Y-%m-%d')
    to_date = datetime.strptime(request.GET.get('toDate'), '%Y-%m-%d')
    date_range = [from_date + timedelta(days=i) for i in range((to_date - from_date).days + 1)]

    planning_data = Planning.objects.filter(
        employee_name__in=emp_names,
        client_name__in=client_names,
        date__range=(from_date, to_date)
    ).values('date', 'employee_name', 'client_name')

    client_employee_pairs = sorted(
        {(entry['client_name'], entry['employee_name']) for entry in planning_data},
        key=lambda x: (x[0], x[1])
    )

    columns = ['Date', 'Day']
    for client, emp in client_employee_pairs:
        columns.append(client)

    plan_lookup = {}
    for entry in planning_data:
        date_str = entry['date'].strftime('%Y-%m-%d')
        key = (entry['client_name'], entry['employee_name'])
        plan_lookup.setdefault(date_str, {})[key] = True

    # Workbook and Sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "UnallocatedData"

    # === IMAGE in A1 ===
    img_path = os.path.join(settings.BASE_DIR, "dashboard", "static", "projectimage", "rspc.png")
    if os.path.exists(img_path):
        img = XLImage(img_path)
        img.width = 250
        img.height = 43
        ws.add_image(img, "A1")
        ws.row_dimensions[1].height = 45
        ws.column_dimensions['A'].width = 55

    # === TITLE in A2 ===
    from_date_dd = from_date.strftime('%d-%m-%Y')
    to_date_dd = to_date.strftime('%d-%m-%Y')
    ws["A2"] = f"Unallocated Resources From {from_date_dd} To {to_date_dd}"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    header_fill = PatternFill(start_color="000C66", end_color="000C66", fill_type="solid")

    # === HEADER in row 3 ===
    for col_num, value in enumerate(columns, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = value
        cell.font = Font(bold=True,color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill

    # === DATA starting from row 4 ===
    for idx, date in enumerate(date_range, start=4):
        date_str = date.strftime('%Y-%m-%d')
        row = [
            date,
            calendar.day_abbr[date.weekday()]  # Short day name
        ]
        for client, emp in client_employee_pairs:
            if not plan_lookup.get(date_str, {}).get((client, emp), False):
                row.append(emp)
            else:
                row.append("")

        
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=idx, column=col_num)
            cell.value = value
            if col_num == 1:  # Date column formatting
                cell.number_format = 'DD-MM-YYYY'
                cell.alignment = Alignment(horizontal='center')

        if calendar.day_abbr[date.weekday()] == "Sun":
            fill = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")  # Coral Red
            for col_num in range(1, len(row) + 1):
                ws.cell(row=idx, column=col_num).fill = fill

    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 12

    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter  # e.g. 'A', 'B', etc.
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        # Adjust column width (+2 for padding)
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    # === Response as Excel File ===
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Unllocated Resources From {from_date_dd} To {to_date_dd}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

